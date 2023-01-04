import streamlit as st
import io
from io import StringIO
from pyxlsb import open_workbook as open_xlsb

import os
# adding a file uploader to accept multiple CSV files
try:
    d='A WEB BASED INTERFACE '
    st.title(d)
    mod =st.number_input("Enter mod value",max_value=10000)
    checkbutton1 = st.button("If you want only a single or few files")
    uploaded_files = st.file_uploader("Please choose a Excel file", type=['xlsx'], accept_multiple_files=True)
    checkbutton2 = st.button("Specify path of the folder for complete conversion")
    vb=st.text_input("Path Input")
    checkbutton = st.button("Compute")
    print(uploaded_files)
except:
    print('Problem in initial output to webpage')

if checkbutton2:
    uploaded_files=os.listdir(vb)
for i in uploaded_files:
	if checkbutton2:
		print(vb)
		i=vb+i
print(vb)
print(uploaded_files)
from datetime import datetime

start_time = datetime.now()
import pandas as pd
import openpyxl
import os

from collections import OrderedDict
import numpy as np
# Help
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import Workbook

fill_cell = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')


def octant_analysis(mod=5000):
    # checking if compute button has been clicked or not
    if checkbutton:
        for file1 in uploaded_files:
            fname = file1.name
            fname=fname.split('.xlsx')[0]


            try:
                df = pd.read_excel(file1)
                print('read file {}',file1)
            except:
                print("File not found")
                return 0
            # importing the file
            pd.options.mode.chained_assignment = None  # default='warn'
            # to remove error that occurs while editing a newly created column
            color = []
            df['U Avg'] = ' '
            df['U Avg'][0] = 0
            # creating a new column to store U_avg

            sum = 0
            for i in df['U']:
                sum += i
            df['U Avg'][0] = sum / df['U'].size
            # calculating the average of the column values for U

            df['V Avg'] = ' '
            sum = 0
            for i in df['V']:
                sum += i
            df['V Avg'][0] = sum / df['V'].size
            # calculating the average of the column values for V

            df['W Avg'] = ' '
            sum = 0
            for i in df['W']:
                sum += i
            df['W Avg'][0] = sum / df['W'].size
            # calculating the average of the column values for W

            uavg = df['U Avg'][0]
            vavg = df['V Avg'][0]
            wavg = df['W Avg'][0]
            df["U'=U -U Avg"] = df['U'] - uavg
            df["V'=V -V Avg"] = df['V'] - vavg
            df["W'=W -W Avg"] = df['W'] - wavg

            # creating U',V',W' columns by subtracting values in U,V,W from their averages

            df['Octant'] = 0
            # new column for octant

            for i in range(df['U'].size):
                val = 1
                if df["U'=U -U Avg"][i] < 0 and df["V'=V -V Avg"][i] > 0:
                    val = 2
                if df["U'=U -U Avg"][i] < 0 and df["V'=V -V Avg"][i] < 0:
                    val = 3
                if df["U'=U -U Avg"][i] > 0 and df["V'=V -V Avg"][i] < 0:
                    val = 4
                if df["W'=W -W Avg"][i] < 0:
                    val = val * -1
                df['Octant'][i] = val

            # as explained in the video , I have assigned octant names for values in U',V',W'
            thisdict = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
            for i in df['Octant']:
                thisdict[i] += 1
            # dictionary made to get the frequency of each octant value (+1,-1,2,-1....)
            ct = 1
            df[' ' * ct] = ' '
            ct += 1
            df[' ' * ct] = ' '
            df[' ' * ct][2] = 'Mod' + ' ' + str(mod)
            ct += 1
            df['Octant ID'] = ' '
            for i in range(1, 5):
                df['+' + str(i)] = ' '
                df['-' + str(i)] = ' '
            for i in range(1, 5):
                df[i] = ' '
                df[-i] = ' '

            # relevant columns have been made

            df['Octant ID'][2] = 'Overall Count'
            df['Octant ID'][1] = 'Octant id'
            df['+1'][2] = thisdict[1]
            df['-1'][2] = thisdict[-1]

            df['+2'][2] = thisdict[2]
            df['-2'][2] = thisdict[-2]
            df['+3'][2] = thisdict[3]
            df['-3'][2] = thisdict[-3]
            df['+4'][2] = thisdict[4]
            df['-4'][2] = thisdict[-4]
            df[+1][1] = 'Rank of 1'
            df[-1][1] = 'Rank of -1'
            df[+2][1] = 'Rank of 2'
            df[-2][1] = 'Rank of -2'
            df[+3][1] = 'Rank of 3'
            df[-3][1] = 'Rank of -3'
            df[+4][1] = 'Rank of 4'
            df[-4][1] = 'Rank of -4'

            k = dict(sorted(thisdict.items(), key=lambda item: item[1]))
            # print(k)
            # frequency values have been assigned to columns using the dictionary created above
            df['Rank1 Octant ID'] = ' '
            df['Rank1 Octant Name'] = ' '
            oct_name_id_mapping = {1: 'Internal outward interaction', -1: 'External outward interaction',
                                   2: 'External Ejection', -2: 'Internal Ejection',
                                   3: 'External inward interaction', -3: 'Internal intward interaction',
                                   4: 'Internal Sweep', -4: 'External Sweep'}
            rank = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
            itr = 8
            for i in k.keys():
                rank[i] = itr
                itr -= 1
            for i in range(1, 5):
                df[i][2] = rank[i]
                df[-i][2] = rank[-i]
                if (rank[i] == 1):
                    df['Rank1 Octant ID'][2] = i
                    df['Rank1 Octant Name'][2] = oct_name_id_mapping[i]
                if (rank[-i] == 1):
                    df['Rank1 Octant ID'][2] = -i
                    df['Rank1 Octant Name'][2] = oct_name_id_mapping[-i]
            # print(rank)

            # different mod need different names
            freq_rank1 = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
            start = 0
            i = 3
            cnt = 0
            while (start < df['U'].size):
                rg = min(mod, df['U'].size - start)
                df['Octant ID'][i] = str(start) + '-' + str(min(start + mod - 1, df['U'].size - 1))
                cnt += 1
                td = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
                for j in range(rg):
                    td[df['Octant'][start + j]] += 1
                df['+1'][i] = td[1]
                df['-1'][i] = td[-1]
                df['+2'][i] = td[2]
                df['-2'][i] = td[-2]
                df['+3'][i] = td[3]
                df['-3'][i] = td[-3]
                df['+4'][i] = td[4]
                df['-4'][i] = td[-4]
                k = dict(sorted(td.items(), key=lambda item: item[1]))
                itr = 8
                for j in k.keys():
                    rank[j] = itr
                    itr -= 1
                for j in range(1, 5):
                    df[j][i] = rank[j]
                    df[-j][i] = rank[-j]
                    if (rank[j] == 1):
                        df['Rank1 Octant ID'][i] = j
                        df['Rank1 Octant Name'][i] = oct_name_id_mapping[j]
                        freq_rank1[j] += 1
                    if (rank[-j] == 1):
                        df['Rank1 Octant ID'][i] = -j
                        df['Rank1 Octant Name'][i] = oct_name_id_mapping[-j]
                        freq_rank1[-j] += 1
                i += 1
                start += mod
            df[4][i + 3] = 'Octant ID'

            df[-4][i + 3] = 'Octant Name'
            df['Rank1 Octant ID'][i + 3] = 'Count of Rank 1 Mod values'
            j = 1
            b = i + 4
            v = b % 2
            for i in range(b, b + 8):
                if (i % 2 == v):
                    df[4][i] = j
                    df[-4][i] = oct_name_id_mapping[j]
                    df['Rank1 Octant ID'][i] = freq_rank1[j]
                else:
                    df[4][i] = -j
                    df[-4][i] = oct_name_id_mapping[-j]
                    df['Rank1 Octant ID'][i] = freq_rank1[-j]
                    j += 1

            # here we have used two loops , first we go from 0->4999,5000->9999 and so on
            # each time we store the values in a dictionary and then assign them to relevant columns
            # for the last time as we don't know how far to go , we use rg(range) = size of column - start of this range
            # as for this example , when we reach 25000, it will take min(5000,29745-25000) , giving us 4745

            #     print(df)
            # final output to csv

            df.rename(columns={'Octant ID': 'Overall Octant Count'}, inplace=True)

            for m in range(1, 5):
                df['+' + str(m)][1] = str(m)
                df['-' + str(m)][1] = '-' + str(m)
                df.rename(columns={m: ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={-m: ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={'+' + str(m): ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={'-' + str(m): ' ' * ct}, inplace=True)
                ct += 1
            df['Rank1 Octant ID'][1] = 'Rank1 Octant ID'
            df['Rank1 Octant Name'][1] = 'Rank1 Octant Name'
            df.rename(columns={'Rank1 Octant ID': ' ' * ct}, inplace=True)
            ct += 1
            df.rename(columns={'Rank1 Octant Name': ' ' * ct}, inplace=True)
            ct += 1
            dp = [[0 for j in range(8)] for i in range(8)]
            for j in range(df['U'].size - 1):
                f = df.at[j, 'Octant']
                s = df.at[j + 1, 'Octant']
                x = 0
                y = 0
                if (f > 0):
                    x = 2 * f - 2
                else:
                    x = 2 * abs(f) - 1
                if (s > 0):
                    y = 2 * s - 2
                else:
                    y = 2 * abs(s) - 1
                dp[x][y] += 1
            df[' ' * ct] = ' '
            df['From'] = ' '
            df['Overall Transition Count'] = ' '
            for l in range(1, 5):
                df['+' + str(l)] = ' '
                df['-' + str(l)] = ' '
            i = 2
            df['From'][i] = 'From'
            df['+1'][i - 2] = 'To'
            df['Overall Transition Count'][i - 1] = 'Octant #'
            k = i
            for j in range(1, 5):
                df['Overall Transition Count'][k] = '+' + str(j)
                k += 1
                df['Overall Transition Count'][k] = '-' + str(j)
                k += 1
            k = i - 1
            for j in range(1, 5):
                df['+' + str(j)][k] = '+' + str(j)
                df['-' + str(j)][k] = '-' + str(j)
            # print(dp)
            for k in range(2, 2 + 8):
                for j in range(35, 35 + 8):
                    df.iat[k, j] = dp[k - 2][j - 35]
            start = 0
            i += 14
            m = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
            # we are going throug each mod range of values and finding transition count for the same 0-5000,5000-10000 etc
            while (start < df['U'].size):

                rg = min(mod, df['U'].size - start)
                df.iat[i - 3, 34] = 'Mod Transition Count'
                df.iat[i - 3, 34] = str(start) + '-' + str(start + rg)
                df.iat[i - 3, 35] = 'To'
                df.iat[i - 2, 34] = 'Octant #'
                df.iat[i - 1, 33] = 'From'
                dp = [[0 for j in range(8)] for i in range(8)]
                # local for each range
                for j in range(start, rg + start - 1):
                    f = df.at[j, 'Octant']
                    s = df.at[j + 1, 'Octant']
                    x = 0
                    y = 0
                    if (f > 0):
                        x = 2 * f - 2
                    else:
                        x = 2 * abs(f) - 1
                    if (s > 0):
                        y = 2 * s - 2
                    else:
                        y = 2 * abs(s) - 1
                    dp[x][y] += 1
                for h in range(1, 5):
                    df['+' + str(h)][i - 2] = '+' + str(h)
                    df['-' + str(h)][i - 2] = '-' + str(h)

                for v in range(i - 1, i + 7):
                    df.at[v, 'Overall Transition Count'] = m[v - i + 1]
                start += mod
                y = i - 1
                # print(dp)
                for k in range(y, y + 8):
                    for j in range(35, 35 + 8):
                        df.iat[k, j] = dp[k - y][j - 35]
                    # print(df.iat[k,j])
                    # print(k,j)
                # inserting for each range
                i += 13
            for m in range(1, 5):
                df['+' + str(m)][1] = str(m)
                df['-' + str(m)][1] = '-' + str(m)
                df.rename(columns={m: ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={-m: ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={'+' + str(m): ' ' * ct}, inplace=True)
                ct += 1
                df.rename(columns={'-' + str(m): ' ' * ct}, inplace=True)
                ct += 1
            df[' ' * ct] = ''
            ct += 1
            storage = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
            loc = {1: [], -1: [], 2: [], -2: [], 3: [], -3: [], 4: [], -4: []}
            freq_of_longestsubsequence = {1: 0, -1: 0, 2: 0, -2: 0, 3: 0, -3: 0, 4: 0, -4: 0}
            c = 0
            for i in range(df['U'].size - 1):
                if (df.at[i, 'Octant'] == df.at[i + 1, 'Octant']):
                    c += 1
                else:
                    storage[df.at[i, 'Octant']] = max(storage[df.at[i, 'Octant']], c + 1)
                    c = 0
            #     print(storage)
            #     print(storage)
            for i in range(df['U'].size - 1):
                if (df.at[i, 'Octant'] == df.at[i + 1, 'Octant']):
                    c += 1
                else:
                    if (c + 1 == storage[df.at[i, 'Octant']]):
                        freq_of_longestsubsequence[df.at[i, 'Octant']] += 1
                        loc[df.at[i, 'Octant']].append(i / 100)
                    c = 0

            df['Longest Subsequence Length'] = ' '
            df['Longest Subsequence Length'][1] = 'Octant #'
            df['v'] = ' '
            df['g'] = ' '
            df['v'][1] = 'Longest Subsequence Length'
            df['g'][1] = 'Count'
            j = 1
            for i in range(0, 8):
                if (i % 2 == 0):
                    df.at[i + 2, 'Longest Subsequence Length'] = '+' + str(j)
                else:
                    df.at[i + 2, 'Longest Subsequence Length'] = '-' + str(j)

                if (i % 2 == 0):
                    df.at[i + 2, 'v'] = storage[j]
                    df.at[i + 2, 'g'] = freq_of_longestsubsequence[j]
                else:
                    df.at[i + 2, 'v'] = storage[0 - j]
                    df.at[i + 2, 'g'] = freq_of_longestsubsequence[0 - j]
                if (i % 2 == 1):
                    j += 1
            df[' ' * ct] = ''
            ct += 1
            timeseries = 0

            # finding longest subsequence for each value in the data
            # df['v1'] = ' '
            # df['g1'] = ' '
            # df['v1'][1] = 'Longest Subsequence Length'
            # df['g1'][1] = 'Count'
            # df['Longest Subsequence Length with Range'][1] = 'Octant #'
            df['Count '] = ' '
            df['Count '][1] = 'Octant #'
            df['Longest Subsequence Length '] = ' '
            df['Longest Subsequence Length '][1] = 'Longest Subsequence Length '
            df['count'] = ' '
            df['count'][1] = 'Count'
            j = 0
            pos = {0: 1, 1: -1, 2: 2, 3: -2, 4: 3, 5: -3, 6: 4, 7: -4}

            # position dictionary to store location for each octant typwe
            for i in range(8):
                df.at[j + 2, 'Count '] = pos[i]

                df.at[j + 2, 'Longest Subsequence Length '] = storage[pos[i]]
                df.at[j + 2, 'count'] = freq_of_longestsubsequence[pos[i]]
                timeseries += freq_of_longestsubsequence[pos[i]]
                j += 1
                df.at[j + 2, 'Count '] = 'Time'
                df.at[j + 2, 'Longest Subsequence Length '] = 'From'
                df.at[j + 2, 'count'] = 'To'
                j += 1
                for k in range(len(loc[pos[i]])):
                    df.at[j + 2, 'Longest Subsequence Length '] = loc[pos[i]][k] - storage[pos[i]] / 100 + 1 / 100
                    df.at[j + 2, 'count'] = loc[pos[i]][k]
                    j += 1
            df.rename(columns={'v': ' ' * ct}, inplace=True)
            ct += 1
            df.rename(columns={'g': ' ' * ct}, inplace=True)
            ct += 1
            df.rename(columns={'From' + str(m): ' ' * ct}, inplace=True)
            ct += 1
            try:
                df.to_excel("temp.xlsx")
            except:
                print('Error while creating output file')
            path = 'temp.xlsx'
            # opening with openpyxl to color the cells and boundary
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # filling color in cell and border
            for x in range(3, 9):
                for y in range(15, 15 + 19):
                    if (sheet_obj.cell(row=x, column=y).value == 1 and x != 3):
                        sheet_obj.cell(row=x, column=y).fill = fill_cell
                    # 	fills yellow color
                    sheet_obj.cell(row=x, column=y).border = thin_border
            # 		fills border
            for x in range(12, 12 + 9):
                for y in range(30, 30 + 3):
                    sheet_obj.cell(row=x, column=y).border = thin_border
            for x in range(3, 3 + 9):
                for y in range(46, 46 + 3):
                    sheet_obj.cell(row=x, column=y).border = thin_border
            for v in range(cnt + 1):
                sd = 3 + v * 13
                for x in range(sd, sd + 9):
                    maxv = -1
                    for y in range(36, 36 + 9):

                        sheet_obj.cell(row=x, column=y).border = thin_border

                        if isinstance(sheet_obj.cell(row=x, column=y).value, str):
                            continue
                        maxv = max(maxv, sheet_obj.cell(row=x, column=y).value)
                    for y in range(36, 36 + 9):
                        if (sheet_obj.cell(row=x, column=y).value == maxv):
                            sheet_obj.cell(row=x, column=y).fill = fill_cell
            for x in range(3, 12 + 2 * timeseries):
                for y in range(50, 50 + 3):
                    sheet_obj.cell(row=x, column=y).border = thin_border
            df.rename(columns={'From': ' ' * ct}, inplace=True)
            ct += 1
            print('rh')
            try:
                wb_obj.save('output.xlsx')
                file_path='output.xlsx'
                try:
                    with open(file_path, 'rb') as my_file:
                        st.download_button(label=fname+'_'+str(mod)+'_'+str(datetime.now())+'.xlsx', data=my_file, file_name=fname+'_'+str(mod)+str(datetime.now())+'.xlsx',
                                           mime='xlsx')
                except:
                    print('error')


            except:
                print("Error in output")




print(mod)
octant_analysis(mod)
